from logging import getLogger
from re import search

from numpy import array, nan
from openpyxl import load_workbook

from pynetix import __project__
from pynetix.models.reaction import Reaction
from pynetix.other.fileformats import Standard, New


class Plate:
    def __init__(self, file, dimensions=(8, 12)) -> None:

        self.metaData = {}
        self.filePath = str(file)
        self.format = None
        self.dimensions = dimensions
        self.reactions = None
        self.times = None
        self.timeUnits = None
        self.results = None

        self._parseFile()

    @property
    def timeLabel(self):
        return f'Time t / {self.timeUnits}'

    def changeMetaData(self, metaData: str, value: str) -> None:
        try:
            wb, ws = self._openWorkbook()
        except:
            getLogger(__project__).error(
                f'Changing meta data failed. File not accessible anymore.', 20)

        if metaData == 'Time':
            pass
        elif metaData == 'Date':
            pass

        ws[self.format.metaData[metaData]['loc']
           ].value = f"{self.format.metaData[metaData]['name']}: {value}"
        wb.save(self.filePath)
        getLogger(__project__).info(
            f"Changed '{metaData}' to '{value}' successfully.")

    def _openWorkbook(self):
        try:
            wb = load_workbook(self.filePath)
            ws = wb.active
            return wb, ws
        except:
            raise ValueError

    def _parseFile(self) -> None:
        _, ws = self._openWorkbook()
        self._checkFileFormat(ws)
        self.metaData = self._parseMetaData(ws)
        valueUnits = self._parseValueUnits(ws)
        self.times = self._parseTime(ws)
        self._parseReactions(ws, valueUnits)

    def _checkFileFormat(self, ws) -> str:
        if ws['A10'].value is None:
            self.format = New
        else:
            self.format = Standard

    def _parseMetaData(self, ws) -> None:
        metaData = {}
        for md, info in self.format.metaData.items():
            raw = ws[info['loc']].value
            reg = fr"^{info['name']}: (.*)$"
            result = search(reg, raw)
            groups = result.groups() if result is not None else ['', ]
            value = groups[0]

            metaData.update({md: value})

        return metaData

    def _parseValueUnits(self, ws) -> str:
        return search(r'as (.*)$', ws[self.format.valueUnits].value).groups()[0]

    def _parseTime(self, ws):
        reg = r'Cycle \d* \((.* h)? ?(.* min)? ?(.* s)?\)'
        delta = 4 + self.dimensions[0]
        i = self.format.timeLoc
        times = []

        while True:
            if ws[f'A{i:d}'].value:
                h, m, s = search(reg, ws[f'A{i:d}'].value).groups()
                m = int(search('(\d*)', m).groups()[0])
                s = 0 if s is None else int(search('(\d*)', s).groups()[0])
                if h is not None:
                    h = int(search('\d*', h).groups()[0])
                    self.timeUnits = 'h'
                else:
                    h = 0
                    self.timeUnits = 's'

                times.append(24*60*h+60*m+s)
                i += delta
            else:
                break

        return array(times)

    def _parseReactions(self, ws, units: str) -> None:
        delta = 4 + self.dimensions[0]
        self.reactions = []

        for row in range(self.dimensions[0]):
            self.reactions.append([])
            for col in range(self.dimensions[1]):
                cycle = 0
                values = []
                while True:
                    value = ws.cell(self.format.startRow+row+cycle,
                                    self.format.startCol+col).value
                    if value:
                        if value in ['overflow']:
                            value = nan
                        values.append(value)
                        cycle += delta
                    else:
                        break
                self.reactions[row].append(
                    Reaction(self, values, row, col, units))
